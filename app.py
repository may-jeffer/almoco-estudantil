from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, Response
import qrcode
from io import BytesIO
import io
import csv
import os
from datetime import datetime, timedelta
import models
import base64
import os

from werkzeug.security import generate_password_hash, check_password_hash
import time

app = Flask(__name__)
# Secret key segura via variável de ambiente ou fallback robusto
app.secret_key = os.environ.get('SECRET_KEY', 'ifap_cantina_ultra_secret_key_2026_@!')

# Inicializa o banco ao rodar
models.init_db()

import os
UPLOAD_FOLDER = os.path.join('static', 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
    
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.context_processor
def inject_config():
    config = models.get_config()
    return dict(global_config=config, tem_permissao=tem_permissao)

def is_logged_in_aluno():
    return 'aluno_id' in session

def is_logged_in_admin():
    return session.get('is_admin') == True

def tem_permissao(p):
    if not is_logged_in_admin(): return False
    perms = session.get('admin_permissoes', [])
    return 'all' in perms or p in perms

def is_admin_mestre():
    return tem_permissao('all')

def datetime_now_str():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
def get_agora():
    return datetime.now()

def date_hoje_str():
    return datetime.now().strftime('%Y-%m-%d')

def pode_reservar(data_cardapio_str, horario_limite_str):
    agora = get_agora()
    data_cardapio = datetime.strptime(data_cardapio_str, '%Y-%m-%d')
    hora, minuto = map(int, horario_limite_str.split(':'))
    
    data_limite = data_cardapio - timedelta(days=1)
    momento_limite = data_limite.replace(hour=hora, minute=minuto, second=0, microsecond=0)
    
    return agora <= momento_limite

# Filtro para Jinja
@app.template_filter('datetimeformat')
def datetimeformat(value, format='%d/%m/%Y'):
    if not value:
        return ''
    try:
        return datetime.strptime(value, '%Y-%m-%d').strftime(format)
    except:
        return value

def sanitize_field(val):
    if not val: return ""
    val = str(val).strip()
    # Prevenir CSV Injection (vulnerabilidade explorada em Excel/Google Sheets)
    if val.startswith(('=', '+', '-', '@')):
        return "'" + val
    return val

# --- Rotas do Aluno ---
@app.route('/')
def index():
    if is_logged_in_aluno():
        return redirect(url_for('aluno_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        cpf = request.form.get('cpf')
        data_nascimento = request.form.get('data_nascimento')
        senha = request.form.get('senha')
        
        config = models.get_config()
        modo_login = config['modo_login_aluno'] if 'modo_login_aluno' in dict(config) else 'DATA_NASC'
        
        with models.closing(models.get_db_connection()) as conn:
            if modo_login == 'DATA_NASC':
                aluno = conn.execute('SELECT * FROM alunos WHERE cpf = ? AND data_nascimento = ?', (cpf, data_nascimento)).fetchone()
                if aluno:
                    session['aluno_id'] = aluno['id']
                    session['aluno_nome'] = aluno['nome']
                    return redirect(url_for('aluno_dashboard'))
                else:
                    flash('CPF ou Data de Nascimento inválidos.', 'error')
            else: # modo_login == 'SENHA'
                aluno = conn.execute('SELECT * FROM alunos WHERE cpf = ?', (cpf,)).fetchone()
                if aluno:
                    senha_valida = False
                    if aluno['senha_hash']:
                        if aluno['senha_hash'].startswith('pbkdf2:sha256:'):
                            senha_valida = check_password_hash(aluno['senha_hash'], senha)
                        else:
                            # Legado SHA-256
                            import hashlib
                            if hashlib.sha256(senha.encode()).hexdigest() == aluno['senha_hash']:
                                senha_valida = True
                                # Migrar para formato seguro
                                novo_hash = generate_password_hash(senha)
                                conn.execute("UPDATE alunos SET senha_hash = ? WHERE id = ?", (novo_hash, aluno['id']))
                                conn.commit()
                    
                    if senha_valida:
                        session['aluno_id'] = aluno['id']
                        session['aluno_nome'] = aluno['nome']
                        return redirect(url_for('aluno_dashboard'))
                    else:
                        # Se não tem senha (1º acesso) ou se errou
                        senhas_provisorias = [
                            aluno['data_nascimento'],
                            aluno['data_nascimento'].replace('-', ''),
                            datetime.strptime(aluno['data_nascimento'], '%Y-%m-%d').strftime('%d%m%Y')
                        ]
                        if senha in senhas_provisorias:
                            session['setup_aluno_id'] = aluno['id']
                            return redirect(url_for('aluno_setup_senha'))
                        
                        time.sleep(1) # Security throttle
                        flash('Senha incorreta ou CPF não cadastrado.', 'error')
                else:
                    time.sleep(1)
                    flash('Estudante não encontrado.', 'error')
                
    return render_template('login.html')

@app.route('/aluno/setup_senha', methods=['GET', 'POST'])
def aluno_setup_senha():
    if 'setup_aluno_id' not in session:
        return redirect(url_for('login'))
        
    if request.method == 'POST':
        email = request.form.get('email')
        nova_senha = request.form.get('nova_senha')
        
        if len(nova_senha) < 6:
            flash('A senha precisa ter no mínimo 6 caracteres.', 'error')
            return redirect(url_for('aluno_setup_senha'))
            
        hash_senha = generate_password_hash(nova_senha)
        
        with models.closing(models.get_db_connection()) as conn:
            conn.execute("UPDATE alunos SET email = ?, senha_hash = ? WHERE id = ?", (email, hash_senha, session['setup_aluno_id']))
            conn.commit()
            
        session.pop('setup_aluno_id', None)
        flash('Cadastro de segurança concluído! Faça login com a sua nova senha.', 'success')
        return redirect(url_for('login'))
        
    return render_template('aluno_setup_senha.html')

@app.route('/esqueci_senha', methods=['GET', 'POST'])
def esqueci_senha():
    if request.method == 'POST':
        cpf = request.form.get('cpf')
        
        with models.closing(models.get_db_connection()) as conn:
            aluno = conn.execute('SELECT * FROM alunos WHERE cpf = ?', (cpf,)).fetchone()
            config = conn.execute('SELECT * FROM configuracoes WHERE id = 1').fetchone()
            
            if not aluno:
                flash('CPF não encontrado no sistema.', 'error')
                return redirect(url_for('esqueci_senha'))
                
            if not aluno['email']:
                flash('Você não cadastrou um e-mail. Solicite o resgate manual na Secretaria do Campus.', 'error')
                return redirect(url_for('esqueci_senha'))
                
            if not config['smtp_ativo'] or not config['smtp_host']:
                flash('Recuperação desativada. O IFAP não habilitou o servidor de E-mail automático.', 'error')
                return redirect(url_for('login'))
                
            # Gerar Hash Token Temporário de Resgate (Validade: 30 Mins)
            import secrets
            token = secrets.token_hex(16)
            expiracao = datetime.now() + timedelta(minutes=30)
            
            conn.execute("UPDATE alunos SET reset_token = ?, reset_expiracao = ? WHERE id = ?", (token, expiracao.strftime('%Y-%m-%d %H:%M:%S'), aluno['id']))
            conn.commit()
            
            # Enviar a Mensagem HTML pura via Servidor Customizado de SMTP
            try:
                import smtplib
                from email.mime.text import MIMEText
                from email.mime.multipart import MIMEMultipart
                
                url_recuperacao = url_for('recuperar_senha', token=token, _external=True)
                
                conteudo_html = f"""
                <div style="font-family: Arial, sans-serif; padding: 20px; border: 1px solid #ddd; max-width: 500px; margin: 0 auto; border-radius: 8px;">
                    <h2 style="color: #2F9E41;">Cantina IFAP</h2>
                    <p>Olá, <strong>{aluno['nome']}</strong>!</p>
                    <p>Recebemos uma solicitação para alterar sua senha no portal da cantina. Clique no botão abaixo para criar uma senha nova.</p>
                    <p style="text-align: center; margin: 30px 0;">
                        <a href="{url_recuperacao}" style="background-color: #CD191E; color: white; padding: 12px 24px; text-decoration: none; border-radius: 4px; font-weight: bold;">Recadastrar Minha Senha</a>
                    </p>
                    <p style="font-size: 0.85rem; color: #555;"><i>Nota: Este link expira em exatamente 30 minutos por motivos de segurança cibernética. Se não foi você quem pediu, apenas apague este e-mail misterioso.</i></p>
                </div>
                """
                
                msg = MIMEMultipart()
                msg['Subject'] = 'Recuperação de Acesso - Cantina IFAP'
                msg['From'] = f"Cantina IFAP <{config['smtp_user']}>"
                msg['To'] = aluno['email']
                msg.attach(MIMEText(conteudo_html, 'html'))
                
                if config['smtp_porta'] == 465:
                    server = smtplib.SMTP_SSL(config['smtp_host'], config['smtp_porta'])
                else:
                    server = smtplib.SMTP(config['smtp_host'], config['smtp_porta'])
                    server.starttls()
                    
                server.login(config['smtp_user'], config['smtp_senha'])
                server.send_message(msg)
                server.quit()
                
                flash(f'Sucesso! Enviamos um link de recuperação para o e-mail cadastrado ({aluno["email"]}).', 'success')
            except Exception as e:
                flash(f'Quebra de Transmissão ou Senha Incorreta no Servidor de Disparo. Avise a direção (Status: {e})', 'error')
                
        return redirect(url_for('login'))
        
    return render_template('esqueci_senha.html')

@app.route('/recuperar_senha/<token>', methods=['GET', 'POST'])
def recuperar_senha(token):
    with models.closing(models.get_db_connection()) as conn:
        aluno = conn.execute("SELECT * FROM alunos WHERE reset_token = ?", (token,)).fetchone()
        
        if not aluno:
            flash('Código de segurança expirado ou já queimado.', 'error')
            return redirect(url_for('login'))
            
        expiracao = datetime.strptime(aluno['reset_expiracao'], '%Y-%m-%d %H:%M:%S')
        if datetime.now() > expiracao:
            flash('Seu Link apodreceu (passaram os 30 minutos vitais). Peça permissão novamente.', 'error')
            return redirect(url_for('esqueci_senha'))
            
        if request.method == 'POST':
            nova_senha = request.form.get('nova_senha')
            if len(nova_senha) < 6:
                flash('Senha muito frágil. Tenha resiliência, ao menos 6 blocos.', 'error')
                return redirect(url_for('recuperar_senha', token=token))
                
            hash_digitado = generate_password_hash(nova_senha)
            # Apagar rastros (Queimar token):
            conn.execute("UPDATE alunos SET senha_hash = ?, reset_token = NULL, reset_expiracao = NULL WHERE id = ?", (hash_digitado, aluno['id']))
            conn.commit()
            
            flash('Seu escudo foi refeito. Senha alterada com sucesso!', 'success')
            return redirect(url_for('login'))
            
    return render_template('recuperar_senha.html', token=token, aluno=aluno)

@app.route('/logout')
def logout():
    session.pop('aluno_id', None)
    session.pop('aluno_nome', None)
    return redirect(url_for('login'))

@app.route('/aluno')
def aluno_dashboard():
    if not is_logged_in_aluno():
        return redirect(url_for('login'))
        
    aluno_id = session['aluno_id']
    hoje_str = date_hoje_str()
    
    config = models.get_config()
    horario_limite = config['horario_limite']
    
    # 1. Almoço de Hoje (se houver, mostrar o código QrCode para ele comer)
    cardapio_hoje = None
    reserva_hoje = None
    qr_code_hoje = None
    
    with models.closing(models.get_db_connection()) as conn:
        cardapio_hoje = conn.execute("SELECT * FROM cardapios WHERE data = ?", (hoje_str,)).fetchone()
        if cardapio_hoje:
            reserva_hoje = conn.execute("SELECT * FROM reservas WHERE aluno_id = ? AND cardapio_id = ? AND status != 'CANCELADA'", (aluno_id, cardapio_hoje['id'])).fetchone()
            if reserva_hoje and reserva_hoje['status'] == 'ATIVA':
                qr = qrcode.QRCode(version=1, box_size=10, border=3)
                qr.add_data(reserva_hoje['codigo_unico'])
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white")
                buffered = BytesIO()
                img.save(buffered, format="PNG")
                qr_code_hoje = base64.b64encode(buffered.getvalue()).decode("utf-8")

    # 2. Próximo Cardápio que AINDA ESTÁ ABERTO para reservas
    proximos_cardapios = models.get_proximos_cardapios(hoje_str)
    cardapio_aberto = None
    reserva_aberto = None
    qr_code_aberto = None
    dentro_do_prazo_aberto = False

    for c in proximos_cardapios:
        if c['data'] != hoje_str and pode_reservar(c['data'], horario_limite):
            cardapio_aberto = c
            dentro_do_prazo_aberto = True
            break
            
    if cardapio_aberto:
        with models.closing(models.get_db_connection()) as conn:
            reserva_aberto = conn.execute("SELECT * FROM reservas WHERE aluno_id = ? AND cardapio_id = ? AND status = 'ATIVA'", (aluno_id, cardapio_aberto['id'])).fetchone()
            if reserva_aberto:
                qr = qrcode.QRCode()
                qr.add_data(reserva_aberto['codigo_unico'])
                qr.make(fit=True)
                img = qr.make_image(fill_color="#4F46E5", back_color="white")
                buffered = BytesIO()
                img.save(buffered, format="PNG")
                qr_code_aberto = base64.b64encode(buffered.getvalue()).decode("utf-8")
                
    # BUSCAR AVISOS EXTRAS
    with models.closing(models.get_db_connection()) as conn:
        avisos = conn.execute("SELECT * FROM avisos ORDER BY id DESC").fetchall()
            
    return render_template(
        'aluno_dashboard.html', 
        cardapio_hoje=cardapio_hoje,
        reserva_hoje=reserva_hoje,
        qr_code_hoje=qr_code_hoje,
        cardapio_aberto=cardapio_aberto,
        reserva_aberto=reserva_aberto,
        qr_code_aberto=qr_code_aberto,
        dentro_do_prazo_aberto=dentro_do_prazo_aberto,
        proximos_cardapios=proximos_cardapios,
        avisos=avisos
    )

@app.route('/aluno/reservar/<int:cardapio_id>', methods=['POST'])
def reservar(cardapio_id):
    if not is_logged_in_aluno(): return redirect(url_for('login'))
    
    aluno_id = session['aluno_id']
    config = models.get_config()
    horario_limite = config['horario_limite']
    
    with models.closing(models.get_db_connection()) as conn:
        cardapio = conn.execute('SELECT * FROM cardapios WHERE id = ?', (cardapio_id,)).fetchone()
        
        if not cardapio:
            flash('Cardápio não encontrado.', 'error')
            return redirect(url_for('aluno_dashboard'))
            
        if not pode_reservar(cardapio['data'], horario_limite):
            flash('O prazo para reservar o almoço desta data já encerrou.', 'error')
            return redirect(url_for('aluno_dashboard'))
            
        existente = conn.execute('SELECT id, status FROM reservas WHERE aluno_id = ? AND cardapio_id = ?', (aluno_id, cardapio_id)).fetchone()
        
        if existente:
            if existente['status'] == 'CANCELADA':
                conn.execute("UPDATE reservas SET status='ATIVA', codigo_unico=? WHERE id = ?", (models.generate_unique_code(), existente['id']))
                conn.commit()
                flash('Reserva reativada com sucesso!', 'success')
            else:
                flash('Você já possui uma reserva.', 'warning')
        else:
            conn.execute(
                "INSERT INTO reservas (aluno_id, cardapio_id, codigo_unico, data_registro) VALUES (?, ?, ?, ?)",
                (aluno_id, cardapio_id, models.generate_unique_code(), datetime_now_str())
            )
            conn.commit()
            flash('Reserva confirmada com sucesso!', 'success')
            
    return redirect(url_for('aluno_dashboard'))

@app.route('/aluno/cancelar/<int:cardapio_id>', methods=['POST'])
def cancelar_reserva(cardapio_id):
    if not is_logged_in_aluno(): return redirect(url_for('login'))
    
    aluno_id = session['aluno_id']
    config = models.get_config()
    horario_limite = config['horario_limite']
    
    with models.closing(models.get_db_connection()) as conn:
        reserva = conn.execute(
            "SELECT r.*, c.data FROM reservas r JOIN cardapios c ON r.cardapio_id = c.id WHERE r.aluno_id = ? AND r.cardapio_id = ?", 
            (aluno_id, cardapio_id)
        ).fetchone()
        
        if reserva and reserva['status'] == 'ATIVA':
            if not pode_reservar(reserva['data'], horario_limite):
                 flash('O prazo para cancelar a reserva já encerrou.', 'error')
            else:
                conn.execute("UPDATE reservas SET status='CANCELADA' WHERE id = ?", (reserva['id'],))
                conn.commit()
                flash('Sua reserva foi cancelada com sucesso.', 'success')
                
    return redirect(url_for('aluno_dashboard'))

# --- Rotas do Administrador ---
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        user = request.form.get('usuario')
        senha = request.form.get('senha')
        with models.closing(models.get_db_connection()) as conn:
            admin = conn.execute("SELECT * FROM administradores WHERE usuario = ?", (user,)).fetchone()
            if admin:
                # Verificar se é um hash do Werkzeug (contém ':') ou senha antiga
                senha_valida = False
                if admin['senha'] and ':' in admin['senha']:
                    try:
                        senha_valida = check_password_hash(admin['senha'], senha)
                    except:
                        senha_valida = False
                
                if not senha_valida:
                    # Tentar formatos legados para migração
                    import hashlib
                    hash_antigo = hashlib.sha256(senha.encode()).hexdigest()
                    if admin['senha'] == hash_antigo or admin['senha'] == senha:
                        senha_valida = True
                        # Migrar para o novo hash seguro imediatamente
                        novo_hash = generate_password_hash(senha)
                        conn.execute("UPDATE administradores SET senha = ? WHERE id = ?", (novo_hash, admin['id']))
                        conn.commit()
                
                if senha_valida:
                    session['is_admin'] = True
                    session['admin_id'] = admin['id']
                    session['admin_usuario'] = admin['usuario']
                    session['admin_perfil'] = admin['perfil'] if 'perfil' in admin.keys() else 'admin_mestre'
                    
                    # Carregar Permissões
                    import json
                    try:
                        session['admin_permissoes'] = json.loads(admin['permissoes'] or '[]')
                    except:
                        session['admin_permissoes'] = []
                        
                    if tem_permissao('fila') and not tem_permissao('all') and len(session['admin_permissoes']) == 1:
                        return redirect(url_for('admin_entrega'))
                        
                    return redirect(url_for('admin_dashboard'))
            
        time.sleep(1) # Prevenir força bruta
        flash('Credenciais inválidas.', 'error')
    return render_template('admin/login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('is_admin', None)
    session.pop('admin_id', None)
    session.pop('admin_usuario', None)
    session.pop('admin_perfil', None)
    session.pop('admin_permissoes', None)
    return redirect(url_for('admin_login'))

@app.route('/admin')
def admin_dashboard():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    
    with models.closing(models.get_db_connection()) as conn:
        total_alunos = conn.execute("SELECT COUNT(*) FROM alunos").fetchone()[0]
        total_turmas = conn.execute("SELECT COUNT(*) FROM turmas").fetchone()[0]
        config = conn.execute("SELECT * FROM configuracoes WHERE id = 1").fetchone()
        
    return render_template('admin/dashboard.html', total_alunos=total_alunos, total_turmas=total_turmas, config=config)

# CRUD de Administradores
@app.route('/admin/administradores', methods=['GET', 'POST'])
def admin_administradores():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('admins'): return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        usuario = request.form.get('usuario')
        senha = request.form.get('senha')
        perfil = request.form.get('perfil', 'operador_fila')
        
        plist = request.form.getlist('permissoes[]')
        import json
        permissoes_json = json.dumps(plist)
        
        # Gerar hash seguro para o administrador
        hash_senha = generate_password_hash(senha)
        
        try:
            with models.closing(models.get_db_connection()) as conn:
                conn.execute("INSERT INTO administradores (usuario, senha, perfil, permissoes) VALUES (?, ?, ?, ?)", (usuario, hash_senha, perfil, permissoes_json))
                conn.commit()
                flash('Credencial adicionada com sucesso!', 'success')
        except Exception as e:
            flash(f'Erro. Usuário já existente ou erro no banco: {e}', 'error')
        return redirect(url_for('admin_administradores'))
        
    with models.closing(models.get_db_connection()) as conn:
        admins = conn.execute("SELECT * FROM administradores ORDER BY id ASC").fetchall()
    return render_template('admin/administradores.html', admins=admins)

@app.route('/admin/administradores/excluir/<int:id>', methods=['POST'])
def admin_administradores_excluir(id):
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('admins'): return redirect(url_for('admin_dashboard'))
    with models.closing(models.get_db_connection()) as conn:
        conn.execute("DELETE FROM administradores WHERE id = ? AND usuario != 'admin'", (id,))
        conn.commit()
        flash('Acesso revogado.', 'success')
    return redirect(url_for('admin_administradores'))

@app.route('/admin/administradores/alterar_senha/<int:id>', methods=['POST'])
def admin_administradores_alterar_senha(id):
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    
    # Pode alterar se for mestre, se tiver permissão de admins ou se for a própria conta
    pode_alterar = session.get('admin_id') == id or tem_permissao('admins')
    if not pode_alterar: return redirect(url_for('admin_dashboard'))
    
    nova_senha = request.form.get('nova_senha')
    if not nova_senha or len(nova_senha) < 4:
        flash('Senha muito curta. Use ao menos 4 caracteres.', 'error')
        return redirect(url_for('admin_administradores'))
        
    hash_senha = generate_password_hash(nova_senha)
    with models.closing(models.get_db_connection()) as conn:
        conn.execute("UPDATE administradores SET senha = ? WHERE id = ?", (hash_senha, id))
        conn.commit()
        flash('Senha do administrador alterada com sucesso!', 'success')
        
    return redirect(url_for('admin_administradores'))

@app.route('/admin/administradores/editar/<int:id>', methods=['POST'])
def admin_administradores_editar(id):
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('admins'): return redirect(url_for('admin_dashboard'))
    
    perfil = request.form.get('perfil', 'operador_fila')
    plist = request.form.getlist('permissoes[]')
    
    import json
    permissoes_json = json.dumps(plist)
    
    with models.closing(models.get_db_connection()) as conn:
        conn.execute("UPDATE administradores SET perfil = ?, permissoes = ? WHERE id = ?", (perfil, permissoes_json, id))
        conn.commit()
        
        # Se o admin editado for o próprio logado, atualizar sessão IMEDIATAMENTE
        if session.get('admin_id') == id:
            import json
            session['admin_perfil'] = perfil
            session['admin_permissoes'] = json.loads(permissoes_json)
            
    flash('Perfil e permissões do administrador atualizados!', 'success')
    return redirect(url_for('admin_administradores'))

# CRUD de Avisos
@app.route('/admin/avisos', methods=['GET', 'POST'])
def admin_avisos():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('avisos'): return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        titulo = request.form.get('titulo')
        mensagem = request.form.get('mensagem')
        tipo = request.form.get('tipo', 'info')
        data_criacao = datetime_now_str()
        
        with models.closing(models.get_db_connection()) as conn:
            conn.execute("INSERT INTO avisos (titulo, mensagem, tipo, data_criacao) VALUES (?, ?, ?, ?)", (titulo, mensagem, tipo, data_criacao))
            conn.commit()
            flash('Aviso publicado para os alunos.', 'success')
        return redirect(url_for('admin_avisos'))
        
    with models.closing(models.get_db_connection()) as conn:
        avisos = conn.execute("SELECT * FROM avisos ORDER BY id DESC").fetchall()
    return render_template('admin/avisos.html', avisos=avisos)

@app.route('/admin/avisos/excluir/<int:id>', methods=['POST'])
def admin_avisos_excluir(id):
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    with models.closing(models.get_db_connection()) as conn:
        conn.execute("DELETE FROM avisos WHERE id = ?", (id,))
        conn.commit()
        flash('Aviso removido do mural!', 'success')
    return redirect(url_for('admin_avisos'))

# CRUD da Turma
@app.route('/admin/turmas', methods=['GET', 'POST'])
def admin_turmas():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('turmas'): return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        if nome:
            with models.closing(models.get_db_connection()) as conn:
                conn.execute("INSERT INTO turmas (nome) VALUES (?)", (nome,))
                conn.commit()
                flash('Turma adicionada', 'success')
        return redirect(url_for('admin_turmas'))
        
    with models.closing(models.get_db_connection()) as conn:
        turmas = conn.execute("SELECT * FROM turmas ORDER BY nome").fetchall()
    return render_template('admin/turmas.html', turmas=turmas)

# CRUD de Alunos
@app.route('/admin/alunos', methods=['GET', 'POST'])
def admin_alunos():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('alunos'): return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        matricula = request.form.get('matricula')
        cpf = request.form.get('cpf')
        data_nascimento = request.form.get('data_nascimento')
        email = request.form.get('email')
        restricoes = request.form.get('restricoes')
        turma_id = request.form.get('turma_id')
        
        try:
            with models.closing(models.get_db_connection()) as conn:
                conn.execute(
                    "INSERT INTO alunos (nome, matricula, cpf, data_nascimento, email, restricoes, turma_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (nome, matricula, cpf, data_nascimento, email, restricoes, turma_id)
                )
                conn.commit()
                flash('Aluno cadastrado!', 'success')
        except Exception:
            flash(f'Erro ao cadastrar. Talvez CPF já exista.', 'error')
        return redirect(url_for('admin_alunos'))
        
    with models.closing(models.get_db_connection()) as conn:
        alunos = conn.execute("SELECT a.*, t.nome as turma_nome FROM alunos a LEFT JOIN turmas t ON a.turma_id = t.id ORDER BY a.nome").fetchall()
        turmas = conn.execute("SELECT * FROM turmas").fetchall()
    return render_template('admin/alunos.html', alunos=alunos, turmas=turmas)

@app.route('/admin/alunos/editar/<int:id>', methods=['POST'])
def admin_alunos_editar(id):
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('alunos'): return redirect(url_for('admin_dashboard'))
    
    nome = request.form.get('nome')
    matricula = request.form.get('matricula')
    cpf = request.form.get('cpf')
    data_nascimento = request.form.get('data_nascimento')
    email = request.form.get('email')
    restricoes = request.form.get('restricoes')
    turma_id = request.form.get('turma_id')
    
    try:
        with models.closing(models.get_db_connection()) as conn:
            conn.execute(
                """UPDATE alunos 
                   SET nome=?, matricula=?, cpf=?, data_nascimento=?, email=?, restricoes=?, turma_id=? 
                   WHERE id=?""",
                (nome, matricula, cpf, data_nascimento, email, restricoes, turma_id, id)
            )
            conn.commit()
            flash('Aluno atualizado com sucesso!', 'success')
    except Exception as e:
        flash('Erro ao atualizar aluno. Verifique cadastros duplicados.', 'error')
        
    return redirect(url_for('admin_alunos'))

@app.route('/admin/alunos/excluir/<int:id>', methods=['POST'])
def admin_alunos_excluir(id):
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('alunos'): return redirect(url_for('admin_dashboard'))
    try:
        with models.closing(models.get_db_connection()) as conn:
            conn.execute("DELETE FROM reservas WHERE aluno_id = ?", (id,))
            conn.execute("DELETE FROM alunos WHERE id = ?", (id,))
            conn.commit()
            flash('Aluno excluído com sucesso!', 'success')
    except Exception as e:
        flash('Erro ao excluir aluno.', 'error')
    return redirect(url_for('admin_alunos'))

@app.route('/admin/alunos/csv_template')
def admin_alunos_csv_template():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    output = "Nome,Matricula,CPF,DataNascimento,TurmaID,Email,Restricoes\nExemplo da Silva,2023001,11122233344,2005-05-20,1,teste@gmail.com,Vegano (opcional)"
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-disposition": "attachment; filename=modelo_importacao_alunos.csv"}
    )
    
@app.route('/admin/alunos/importar', methods=['POST'])
def admin_alunos_importar():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    
    if 'arquivo_csv' not in request.files:
        flash('Nenhum arquivo enviado.', 'error')
        return redirect(url_for('admin_alunos'))
        
    file = request.files['arquivo_csv']
    if file.filename == '':
        flash('Nenhum arquivo selecionado.', 'error')
        return redirect(url_for('admin_alunos'))
        
    try:
        stream = io.StringIO(file.stream.read().decode("utf-8-sig"), newline=None)
        csv_input = csv.reader(stream, delimiter=';')
        
        # Faz uma leitura caso use vírgula e retenta se apenas puxar 1 campo na row
        first_row = next(csv_input, None)
        delimiter_used = ';'
        if first_row and len(first_row) == 1 and ',' in first_row[0]:
            delimiter_used = ','
            
        # Re-set flow and re-read correctly
        stream.seek(0)
        csv_input = csv.reader(stream, delimiter=delimiter_used)
        next(csv_input, None) # pula o cabecalho de fato
        
        adicionados = 0
        erros = 0
        
        with models.closing(models.get_db_connection()) as conn:
            for row in csv_input:
                if len(row) >= 5:
                    nome = sanitize_field(row[0])
                    matricula = sanitize_field(row[1])
                    cpf = sanitize_field(row[2])
                    data_nascimento = sanitize_field(row[3])
                    turma_id = sanitize_field(row[4])
                    email = sanitize_field(row[5]) if len(row) > 5 else ''
                    restricoes = sanitize_field(row[6]) if len(row) > 6 else ''
                    
                    if not nome or not cpf:
                        continue
                        
                    try:
                        conn.execute(
                            "INSERT INTO alunos (nome, matricula, cpf, data_nascimento, email, restricoes, turma_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
                            (nome, matricula, cpf, data_nascimento, email, restricoes, turma_id)
                        )
                        adicionados += 1
                    except Exception as e:
                        erros += 1
            conn.commit()
            
        flash(f'Importação concluída! Sucesso: {adicionados} alunos. Erros ignorados: {erros}.', 'success')
        
    except Exception as e:
        flash(f'Erro inexperado lendo arquivo CSV: {str(e)}', 'error')
        
    return redirect(url_for('admin_alunos'))

@app.route('/admin/alunos/reset_senha/<int:id>', methods=['POST'])
def admin_alunos_resetar_senha(id):
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('alunos'): return redirect(url_for('admin_dashboard'))
    
    with models.closing(models.get_db_connection()) as conn:
        conn.execute("UPDATE alunos SET senha_hash = NULL WHERE id = ?", (id,))
        conn.commit()
        flash('Senha do aluno resetada! O próximo acesso voltará a ser feito pela Data de Nascimento provisoriamente.', 'success')
    return redirect(url_for('admin_alunos'))

@app.route('/admin/alunos/suap_sync', methods=['POST'])
def admin_alunos_suap_sync():
    """
    Rota Placeholder para a futura integração direta via API do SUAP.
    O desenvolvedor deverá substituir o MOCK pela biblioteca 'requests'.
    """
    import json
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('alunos'): return redirect(url_for('admin_dashboard'))
    
    api_url = request.form.get('api_url')
    token = request.form.get('token')
    turma_id = request.form.get('turma_id')
    
    # ESPAÇO PARA IMPLEMENTAÇÃO FUTURA:
    # import requests
    # try:
    #     response = requests.get(api_url, headers={"Authorization": f"Bearer {token}"})
    #     response.raise_for_status()
    #     alunos_suap = response.json()
    # except Exception as e:
    #     flash(f"Erro Real API: {e}", "error")
    #     return redirect(url_for('admin_alunos'))
    
    # Mocking (Dados Simulados)
    mock_json = '''
    [
        {"nome": "Maria Emília (Teste SUAP)", "matricula": "202611SUAP", "cpf": "234.345.567-89", "data_nascimento": "2005-08-20"},
        {"nome": "Felipe Souza (Teste SUAP)", "matricula": "202612SUAP", "cpf": "098.876.654-32", "data_nascimento": "2006-12-15"}
    ]
    '''
    try:
        alunos_suap = json.loads(mock_json)
        adicionados = 0
        with models.closing(models.get_db_connection()) as conn:
            for asuap in alunos_suap:
                existe = conn.execute("SELECT id FROM alunos WHERE matricula = ? OR cpf = ?", (asuap['matricula'], asuap['cpf'])).fetchone()
                if not existe:
                    conn.execute(
                        "INSERT INTO alunos (nome, matricula, cpf, data_nascimento, restricoes, turma_id) VALUES (?, ?, ?, ?, ?, ?)",
                        (asuap['nome'], asuap['matricula'], asuap['cpf'], asuap['data_nascimento'], '', turma_id)
                    )
                    adicionados += 1
            conn.commit()
            
        flash(f'Simulação SUAP Executada! O código-base funcionou e {adicionados} aluno(s) injetado(s) com sucesso.', 'success')
    except Exception as e:
        flash(f'Erro na Simulação SUAP: {str(e)}', 'error')
        
    return redirect(url_for('admin_alunos'))

# CRUD Cardapios
@app.route('/admin/cardapios', methods=['GET', 'POST'])
def admin_cardapios():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('cardapios'): return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        data = request.form.get('data')
        descricao = request.form.get('descricao')
        try:
            with models.closing(models.get_db_connection()) as conn:
                conn.execute("INSERT INTO cardapios (data, descricao) VALUES (?, ?)", (data, descricao))
                conn.commit()
                flash('Cardápio adicionado!', 'success')
        except Exception:
            flash(f'Erro. Data possivelmente já existe.', 'error')
        return redirect(url_for('admin_cardapios'))
        
    with models.closing(models.get_db_connection()) as conn:
        cardapios = conn.execute("SELECT * FROM cardapios ORDER BY data DESC").fetchall()
    return render_template('admin/cardapios.html', cardapios=cardapios)

# Excluir Cardapio
@app.route('/admin/cardapios/excluir/<int:id>', methods=['POST'])
def admin_cardapios_excluir(id):
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('cardapios'): return redirect(url_for('admin_dashboard'))
    with models.closing(models.get_db_connection()) as conn:
        r = conn.execute("SELECT COUNT(*) FROM reservas WHERE cardapio_id = ? AND status IN ('ATIVA', 'CONSUMIDA')", (id,)).fetchone()[0]
        if r > 0:
            flash('Impossível excluir. Há reservas ativas ou entregues para este cardápio.', 'error')
        else:
            conn.execute("DELETE FROM reservas WHERE cardapio_id = ?", (id,))
            conn.execute("DELETE FROM cardapios WHERE id = ?", (id,))
            conn.commit()
            flash('Cardápio excluído com sucesso.', 'success')
    return redirect(url_for('admin_cardapios'))

# Editar Cardapio
@app.route('/admin/cardapios/editar/<int:id>', methods=['POST'])
def admin_cardapios_editar(id):
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('cardapios'): return redirect(url_for('admin_dashboard'))
    descricao = request.form.get('descricao')
    with models.closing(models.get_db_connection()) as conn:
        conn.execute("UPDATE cardapios SET descricao = ? WHERE id = ?", (descricao, id))
        conn.commit()
        flash('Cardápio atualizado com sucesso.', 'success')
    return redirect(url_for('admin_cardapios'))

# Configurações
@app.route('/admin/configuracoes', methods=['POST'])
def save_configuracoes():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('config'): return redirect(url_for('admin_dashboard'))
    horario_limite = request.form.get('horario_limite')
    nome_sistema = request.form.get('nome_sistema', 'Cantina Estudantil')
    sigla_instituicao = request.form.get('sigla_instituicao', 'IFAP')
    modo_login_aluno = request.form.get('modo_login_aluno', 'DATA_NASC')
    
    with models.closing(models.get_db_connection()) as conn:
        conn.execute("UPDATE configuracoes SET horario_limite = ?, nome_sistema = ?, sigla_instituicao = ?, modo_login_aluno = ? WHERE id = 1", (horario_limite, nome_sistema, sigla_instituicao, modo_login_aluno))
        conn.commit()
        flash('Configurações gerais salvas com sucesso.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/configuracoes/logo', methods=['POST'])
def save_logo():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('config'): return redirect(url_for('admin_dashboard'))
    
    file = request.files.get('logo_file')
    if file and file.filename != '':
        filename = "logo_instituicao" + os.path.splitext(file.filename)[1]
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        logo_url = '/static/uploads/' + filename
        with models.closing(models.get_db_connection()) as conn:
            conn.execute("UPDATE configuracoes SET logo_path = ? WHERE id = 1", (logo_url,))
            conn.commit()
            flash('Logomarca da Instituição atualizada com sucesso.', 'success')
    else:
         flash('Nenhum arquivo enviado.', 'error')
         
    return redirect(url_for('admin_dashboard'))

# SMTP Configurações
@app.route('/admin/configuracoes/smtp', methods=['POST'])
def save_configuracoes_smtp():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('smtp'): return redirect(url_for('admin_dashboard'))
    
    smtp_ativo = 1 if request.form.get('smtp_ativo') == '1' else 0
    smtp_host = request.form.get('smtp_host')
    smtp_porta = request.form.get('smtp_porta')
    smtp_user = request.form.get('smtp_user')
    smtp_senha = request.form.get('smtp_senha')
    
    with models.closing(models.get_db_connection()) as conn:
        if smtp_senha:
            conn.execute("UPDATE configuracoes SET smtp_ativo=?, smtp_host=?, smtp_porta=?, smtp_user=?, smtp_senha=? WHERE id=1", 
                         (smtp_ativo, smtp_host, smtp_porta, smtp_user, smtp_senha))
        else:
            conn.execute("UPDATE configuracoes SET smtp_ativo=?, smtp_host=?, smtp_porta=?, smtp_user=? WHERE id=1", 
                         (smtp_ativo, smtp_host, smtp_porta, smtp_user))
        conn.commit()
        flash('Configuração do Servidor SMTP salva com sucesso.', 'success')
        
    return redirect(url_for('admin_dashboard'))

# Modulo Fila (Entrega)
@app.route('/admin/entrega')
def admin_entrega():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('fila'): return redirect(url_for('admin_dashboard'))
    hoje_str = date_hoje_str()
    
    with models.closing(models.get_db_connection()) as conn:
        cardapio_hoje = conn.execute("SELECT * FROM cardapios WHERE data = ?", (hoje_str,)).fetchone()
        reservas_hoje = []
        if cardapio_hoje:
            reservas_hoje = conn.execute("""
                SELECT r.id, r.codigo_unico, r.status, a.nome, a.matricula
                FROM reservas r
                JOIN alunos a ON r.aluno_id = a.id
                WHERE r.cardapio_id = ? AND r.status != 'CANCELADA'
                ORDER BY a.nome ASC
            """, (cardapio_hoje['id'],)).fetchall()
        
    https_alert = not request.is_secure
        
    return render_template('admin/entrega.html', cardapio_hoje=cardapio_hoje, reservas_hoje=reservas_hoje, https_alert=https_alert)

@app.route('/admin/entrega/baixar', methods=['POST'])
def baixar_reserva():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('fila'): return redirect(url_for('admin_dashboard'))
    
    codigo = request.form.get('codigo_unico')
    cardapio_id = request.form.get('cardapio_id')
    
    if not codigo or not cardapio_id:
        flash('Código ou cardápio inválido', 'error')
        return redirect(url_for('admin_entrega'))
        
    with models.closing(models.get_db_connection()) as conn:
        reserva = conn.execute(
            "SELECT r.*, a.nome as aluno_nome FROM reservas r JOIN alunos a ON r.aluno_id = a.id WHERE r.codigo_unico = ? AND r.cardapio_id = ?", 
            (codigo, cardapio_id)
        ).fetchone()
        
        if not reserva:
            flash('Reserva não encontrada para este código!', 'error')
        elif reserva['status'] == 'CONSUMIDA':
            flash(f"Refeição já entregue para {reserva['aluno_nome']}!", 'warning')
        elif reserva['status'] == 'CANCELADA':
            flash(f"A reserva de {reserva['aluno_nome']} foi cancelada! Recuse a entrega.", 'error')
        else:
            conn.execute("UPDATE reservas SET status='CONSUMIDA' WHERE id = ?", (reserva['id'],))
            conn.commit()
            flash(f"Refeição liberada: {reserva['aluno_nome']}", 'success')
            
    return redirect(url_for('admin_entrega'))

# API para Baixa Rápida via AJAX (Evita sumiço da câmera)
@app.route('/admin/api/baixar', methods=['POST'])
def api_baixar_reserva():
    if not is_logged_in_admin(): 
        return {"success": False, "message": "Não autorizado"}, 401
    if not tem_permissao('fila'):
        return {"success": False, "message": "Sem permissão de acesso à fila", "type": "error"}, 403
    
    data = request.get_json()
    codigo = data.get('codigo_unico')
    cardapio_id = data.get('cardapio_id')
    
    if not codigo or not cardapio_id:
        return {"success": False, "message": "Dados inválidos"}, 400
        
    with models.closing(models.get_db_connection()) as conn:
        reserva = conn.execute(
            "SELECT r.*, a.nome as aluno_nome FROM reservas r JOIN alunos a ON r.aluno_id = a.id WHERE r.codigo_unico = ? AND r.cardapio_id = ?", 
            (codigo, cardapio_id)
        ).fetchone()
        
        if not reserva:
            return {"success": False, "message": "Reserva não encontrada!", "type": "error"}
        elif reserva['status'] == 'CONSUMIDA':
            return {"success": False, "message": f"Refeição já entregue para {reserva['aluno_nome']}!", "type": "warning"}
        elif reserva['status'] == 'CANCELADA':
            return {"success": False, "message": f"A reserva de {reserva['aluno_nome']} foi cancelada!", "type": "error"}
        else:
            conn.execute("UPDATE reservas SET status='CONSUMIDA' WHERE id = ?", (reserva['id'],))
            conn.commit()
            return {"success": True, "message": f"Refeição liberada: {reserva['aluno_nome']}", "aluno": reserva['aluno_nome']}

@app.route('/admin/api/adicionar_extra', methods=['POST'])
def api_adicionar_extra():
    if not is_logged_in_admin(): 
        return {"success": False, "message": "Não autorizado"}, 401
    if not tem_permissao('fila'):
        return {"success": False, "message": "Sem permissão de acesso à fila", "type": "error"}, 403
    
    data = request.get_json()
    busca = data.get('aluno_busca') # Pode ser matrícula ou CPF
    cardapio_id = data.get('cardapio_id')
    
    if not busca or not cardapio_id:
        return {"success": False, "message": "Dados inválidos"}, 400
        
    with models.closing(models.get_db_connection()) as conn:
        aluno = conn.execute("SELECT id, nome FROM alunos WHERE matricula = ? OR cpf = ?", (busca, busca)).fetchone()
        
        if not aluno:
            return {"success": False, "message": "Estudante não encontrado no sistema.", "type": "error"}
            
        # Verificar se já comeu hoje (por reserva normal ou extra)
        ja_comeu = conn.execute("SELECT id, status, tipo_consumo FROM reservas WHERE aluno_id = ? AND cardapio_id = ? AND status = 'CONSUMIDA'", (aluno['id'], cardapio_id)).fetchone()
        if ja_comeu:
            return {"success": False, "message": f"Atenção: {aluno['nome']} já consumiu refeição hoje!", "type": "warning"}
            
        # Verificar se ele tem uma reserva ATIVA e dar baixa normal para evitar duplicidade
        reserva_ativa = conn.execute("SELECT id FROM reservas WHERE aluno_id = ? AND cardapio_id = ? AND status = 'ATIVA'", (aluno['id'], cardapio_id)).fetchone()
        if reserva_ativa:
            conn.execute("UPDATE reservas SET status='CONSUMIDA' WHERE id = ?", (reserva_ativa['id'],))
            conn.commit()
            return {"success": True, "message": f"Foi dada baixa na reserva PADRÃO de {aluno['nome']}.", "aluno": aluno['nome']}
        
        # Se não comeu e não tem reserva ativa, insere a sobra
        codigo = models.generate_unique_code()
        conn.execute(
            "INSERT INTO reservas (aluno_id, cardapio_id, status, codigo_unico, data_registro, tipo_consumo) VALUES (?, ?, 'CONSUMIDA', ?, ?, 'EXTRA')",
            (aluno['id'], cardapio_id, codigo, datetime_now_str())
        )
        conn.commit()
        
        return {"success": True, "message": f"Sobra entregue: {aluno['nome']} (Marcado como Extra)", "aluno": aluno['nome']}

# Modulo Relatórios
@app.route('/admin/relatorios')
def admin_relatorios():
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('relatorios'): return redirect(url_for('admin_dashboard'))
    with models.closing(models.get_db_connection()) as conn:
        # SQL para buscar cardápios e as contagens de reservas vinculadas a eles
        cardapios = conn.execute("""
            SELECT c.*, 
                   (SELECT COUNT(*) FROM reservas r WHERE r.cardapio_id = c.id AND r.status != 'CANCELADA') as total_reservas,
                   (SELECT COUNT(*) FROM reservas r WHERE r.cardapio_id = c.id AND r.status = 'CONSUMIDA') as total_consumidas
            FROM cardapios c 
            ORDER BY c.data DESC
        """).fetchall()
    return render_template('admin/relatorios_filtro.html', cardapios=cardapios)

@app.route('/admin/relatorios/<int:cardapio_id>')
def admin_relatorio_dia(cardapio_id):
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('relatorios'): return redirect(url_for('admin_dashboard'))
    
    turma_id = request.args.get('turma_id', type=int)
    status_filter = request.args.get('status')
    
    with models.closing(models.get_db_connection()) as conn:
        cardapio = conn.execute("SELECT * FROM cardapios WHERE id = ?", (cardapio_id,)).fetchone()
        
        query = """
            SELECT a.nome, a.matricula, a.restricoes, t.nome as turma_nome, r.status, r.tipo_consumo, t.id as turma_id
            FROM reservas r
            JOIN alunos a ON r.aluno_id = a.id
            LEFT JOIN turmas t ON a.turma_id = t.id
            WHERE r.cardapio_id = ? AND r.status != 'CANCELADA'
        """
        params = [cardapio_id]
        
        if turma_id:
            query += " AND t.id = ?"
            params.append(turma_id)
        if status_filter:
            query += " AND r.status = ?"
            params.append(status_filter)
            
        query += " ORDER BY a.nome ASC"
        reservas = conn.execute(query, params).fetchall()
        
        # Dados para Filtros e Resumo
        turmas = conn.execute("SELECT * FROM turmas ORDER BY nome ASC").fetchall()
        
        # Cálculo de Resumo por Turma (Consolidado)
        resumo_query = """
            SELECT t.nome as turma_nome, 
                   SUM(CASE WHEN r.tipo_consumo = 'NORMAL' OR r.tipo_consumo IS NULL THEN 1 ELSE 0 END) as total,
                   SUM(CASE WHEN r.status = 'CONSUMIDA' AND (r.tipo_consumo = 'NORMAL' OR r.tipo_consumo IS NULL) THEN 1 ELSE 0 END) as consumidas_normais,
                   SUM(CASE WHEN r.status = 'CONSUMIDA' AND r.tipo_consumo = 'EXTRA' THEN 1 ELSE 0 END) as consumidas_extras
            FROM reservas r
            JOIN alunos a ON r.aluno_id = a.id
            LEFT JOIN turmas t ON a.turma_id = t.id
            WHERE r.cardapio_id = ? AND r.status != 'CANCELADA'
            GROUP BY t.id
            ORDER BY t.nome ASC
        """
        resumo_turmas = conn.execute(resumo_query, (cardapio_id,)).fetchall()
        
        total_ativas = len([r for r in reservas if r['status'] == 'ATIVA'])
        consumidas_normal = len([r for r in reservas if r['status'] == 'CONSUMIDA' and (r['tipo_consumo'] == 'NORMAL' or not r['tipo_consumo'])])
        consumidas_extra = len([r for r in reservas if r['status'] == 'CONSUMIDA' and r['tipo_consumo'] == 'EXTRA'])
        
        total_consumidas = consumidas_normal + consumidas_extra
        total_geral = total_ativas + consumidas_normal
        
    return render_template('admin/relatorio_dia.html', 
                           cardapio=cardapio, 
                           reservas=reservas, 
                           total_geral=total_geral, 
                           total_consumidas=total_consumidas,
                           consumidas_normal=consumidas_normal,
                           consumidas_extra=consumidas_extra,
                           turmas=turmas,
                           resumo_turmas=resumo_turmas,
                           filtro_turma=turma_id,
                           filtro_status=status_filter)

@app.route('/admin/relatorios/<int:cardapio_id>/excel')
def admin_relatorio_excel(cardapio_id):
    if not is_logged_in_admin(): return redirect(url_for('admin_login'))
    if not tem_permissao('relatorios'): return redirect(url_for('admin_dashboard'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    with models.closing(models.get_db_connection()) as conn:
        cardapio = conn.execute("SELECT * FROM cardapios WHERE id = ?", (cardapio_id,)).fetchone()
        reservas = conn.execute("""
            SELECT a.nome, a.matricula, a.restricoes, t.nome as turma_nome, r.status, r.tipo_consumo
            FROM reservas r
            JOIN alunos a ON r.aluno_id = a.id
            LEFT JOIN turmas t ON a.turma_id = t.id
            WHERE r.cardapio_id = ? AND r.status != 'CANCELADA'
            ORDER BY t.nome, a.nome ASC
        """, (cardapio_id,)).fetchall()
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio de Refeicoes"
    
    # Cabeçalho
    config = models.get_config()
    ws['A1'] = f"RELATÓRIO DE CONSUMO - {config['sigla_instituicao']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    ws['A2'] = f"Data: {datetime.strptime(cardapio['data'], '%Y-%m-%d').strftime('%d/%m/%Y')}"
    ws['A3'] = f"Cardápio: {cardapio['descricao']}"
    
    # Tabela
    headers = ["Estudante", "Matrícula", "Turma", "Situação", "Tipo", "Obser. Dieta"]
    ws.append([]) # Espaço
    ws.append(headers)
    
    header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    for cell in ws[5]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        
    for r in reservas:
        status_txt = "Entregue" if r['status'] == 'CONSUMIDA' else "Pendente"
        tipo_txt = "Sobra/Extra" if r['tipo_consumo'] == 'EXTRA' else "Reserva Normal"
        ws.append([r['nome'], r['matricula'], r['turma_nome'] or "---", status_txt, tipo_txt, r['restricoes'] or "---"])
        
    # Ajuste de largura
    from openpyxl.utils import get_column_letter
    for i, column_cells in enumerate(ws.columns, 1):
        # Ignorar o cálculo para a primeira linha se ela estiver mesclada, ou apenas pegar o maior valor
        length = 0
        for cell in column_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > length: length = len(val)
            except:
                pass
        ws.column_dimensions[get_column_letter(i)].width = length + 5
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Relatorio_Refeicao_{cardapio['data']}.xlsx"
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": f"attachment; filename={filename}"}
    )

if __name__ == '__main__':
    # Para testes mobile com iPhone, usamos um contexto SSL adhoc
    # (Requer pacote pyopenssl: pip install pyopenssl)
    try:
        app.run(host='0.0.0.0', port=5000, ssl_context='adhoc', debug=True)
    except Exception as e:
        print("Aviso: sem pyopenssl rodando em HTTP (Iphones bloqueiam camera aqui)")
        app.run(host='0.0.0.0', port=5000, debug=True)
